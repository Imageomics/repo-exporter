import os
import pandas as pd
from github import Github, GithubException, Auth
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time

# Config
ORG_NAME = "Imageomics"
OUTPUT_FILE = f"{ORG_NAME}_repo_info.xlsx"

# Helper Functions

def get_file(repo, *paths: str):
    for path in paths:
        try:
            return repo.get_contents(path)
        except GithubException:
            continue
    return "N/A"
    
def get_readme(repo):
    try:
        return repo.get_readme()
    except GithubException:
        return "N/A"

def get_license(repo):
    try:
        return repo.get_license()
    except GithubException:
        return "N/A"
        
def get_num_branches(repo):
    try:
        return repo.get_branches().totalCount
    except:
        return "N/A"
    
def get_repo_creator(repo):
    try:
        commits = repo.get_commits()
        first_commit = commits.reversed[0]
        author = first_commit.author
        return f"{author.name} ({author.login})" if author else "Unknown"
    except Exception:
        return "Unknown"
    
def get_top_contributors(repo, top_n: int = 4) -> str:
    try:
        # Keep fetching until stats are ready
        stats = None
        for _ in range(5):  # Try up to 5 times
            stats = repo.get_stats_contributors()
            if stats:
                break
            time.sleep(10)  # Wait 10 seconds between tries

        if not stats:
            return "Unknown"
        
        contributors = []
        for contributor in stats:
            total_additions = sum(week.a for week in contributor.weeks)
            total_deletions = sum(week.d for week in contributor.weeks)
            total_changes = total_additions + total_deletions
            contributors.append((contributor.author.login if contributor.author else "Unknown", total_changes))

        top_n_contributors = sorted(contributors, key=lambda x: x[1], reverse=True)[:top_n] # sort and take the top N results
        return ", ".join([f"{name} ({changes} changes)" for name, changes in top_n_contributors])
    except Exception:
        return "Unknown"
    
def get_repo_info(repo):
    return {
        "Name": repo.name,
        "Description": repo.description or "N/A",
        "Date Created": repo.created_at.strftime("%Y-%m-%d"),
        "Last Updated": repo.updated_at.strftime("%Y-%m-%d"),
        "Created By": get_repo_creator(repo),
        "Top 4 Contributors": get_top_contributors(repo, 4),
        "Stars": repo.stargazers_count,
        "Has README": "Yes" if get_readme(repo) != "N/A" else "No",
        "Has License": "Yes" if get_license(repo) != "N/A" else "No",
        "Has .gitignore": "Yes" if get_file(repo, ".gitignore") != "N/A" else "No",
        "Has CITATION.cff": "Yes" if get_file(repo, "CITATION.cff") != "N/A" else "No",
        "Has Package Requirements": "Yes" if get_file(repo, "requirements.txt", "environment.yaml", "environment.yml") != "N/A" else "No",
        "Branches": get_num_branches(repo),
    }

def add_excel_color_coding(file_path: str) -> None:
    wb = load_workbook(file_path)
    ws = wb.active

    red_hex_code = "FFC7CE"

    red_fill = PatternFill(start_color=red_hex_code, end_color=red_hex_code, fill_type="solid")

    for row in ws.iter_rows(min_row=2): # skip header
        for cell in row:
            value = str(cell.value).strip().lower()
            if value == "no":
                cell.fill = red_fill

    wb.save(file_path)
    print(f"Color coding applied to {file_path}")
# --------

def main():
    TOKEN = os.getenv("GITHUB_TOKEN") or input("Enter your GitHub token: ").strip()

    gh = Github(auth=Auth.Token(TOKEN))
    try:
        org = gh.get_organization(ORG_NAME)
    except Exception as e:
        print(f"ERROR: Could not access org: \"{ORG_NAME}\"")
        return
    
    print("")
    print(f"Fetching repositories from organization: {ORG_NAME}")
    print("")
    print("----------------")

    repos = org.get_repos()
    data = []

    for repo in repos:
        try:
            info = get_repo_info(repo)
            data.append(info)
            print(f"Fetched info for /{repo.name} repo")
        except Exception as e:
            print(f"ERROR: Cannot fetch /{repo.name} info, due to {type(e).__name__}: {e}. Skipping...")

    if not data:
        print("ERROR: No data collected")
        return
    
    print("----------------")
    print("")

    df = pd.DataFrame(data)
    df.sort_values(by="Name", inplace=True)
    df.to_excel(OUTPUT_FILE, index=False)
    print(f"Finished fetching info for {len(df)} repositories from {ORG_NAME} organization to {OUTPUT_FILE}")
    
    add_excel_color_coding(OUTPUT_FILE)

if __name__ == "__main__":
    main()